from datetime import datetime
from io import BytesIO
from pathlib import Path

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, current_app, jsonify, abort
from sqlalchemy.exc import SQLAlchemyError

from meu_app.models import Pedido, Cliente, Produto, ItemPedido, db
from meu_app.pedidos.services import PedidoService
from meu_app.decorators import login_obrigatorio, permissao_necessaria

pedidos_bp = Blueprint('pedidos', __name__, url_prefix='/pedidos')


@pedidos_bp.route('/', methods=['GET'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def listar_pedidos():
    """Lista todos os pedidos com filtros e ordenação"""
    try:
        filtro_status = request.args.get('filtro', 'todos')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        mes_ano = request.args.get('mes_ano')
        ordenar_por = request.args.get('sort', 'data')
        direcao = request.args.get('direction', 'desc')

        if mes_ano:
            try:
                ano, mes = map(int, mes_ano.split('-'))
                data_inicio_dt = datetime(ano, mes, 1)
                # Calcula o último dia do mês
                if mes == 12:
                    data_fim_dt = datetime(ano + 1, 1, 1)
                else:
                    data_fim_dt = datetime(ano, mes + 1, 1)
                
                data_inicio = data_inicio_dt.strftime('%Y-%m-%d')
                data_fim = data_fim_dt.strftime('%Y-%m-%d')

            except (ValueError, TypeError):
                flash('Formato de mês/ano inválido. Use YYYY-MM.', 'warning')
                mes_ano = '' # Limpa para não ser enviado ao template

        # Validar parâmetros de ordenação
        campos_validos = ['id', 'cliente', 'data', 'valor', 'status']
        if ordenar_por not in campos_validos:
            ordenar_por = 'data'
        
        if direcao not in ['asc', 'desc']:
            direcao = 'desc'
        
        pedidos = PedidoService.listar_pedidos(filtro_status, data_inicio, data_fim, ordenar_por, direcao)
        total_pedidos_filtrados = sum((item.get('total_venda') or 0) for item in pedidos)
        
        current_app.logger.info(f"Listagem de pedidos acessada por {session.get('usuario_nome', 'N/A')} - Ordenado por: {ordenar_por} ({direcao})")
        
        return render_template(
            'listar_pedidos.html',
            pedidos=pedidos,
            filtro=filtro_status,
            data_inicio=data_inicio or '',
            data_fim=data_fim or '',
            mes_ano=mes_ano or '',
            current_sort=ordenar_por,
            current_direction=direcao,
            total_pedidos_filtrados=total_pedidos_filtrados
        )
        
    except Exception as e:
        current_app.logger.error(f"Erro ao listar pedidos: {str(e)}")
        flash(f"Erro ao carregar pedidos: {str(e)}", 'error')
        return render_template('listar_pedidos.html', pedidos=[], filtro='todos', total_pedidos_filtrados=0)

@pedidos_bp.route('/importar-historico', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def importar_historico():
    """Importa pedidos históricos; não há página separada, apenas upload direto."""
    if 'arquivo' not in request.files or not request.files['arquivo'].filename:
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('pedidos.listar_pedidos'))
    arquivo = request.files['arquivo']
    extensao = arquivo.filename.rsplit('.', 1)[1].lower() if '.' in arquivo.filename else ''
    if extensao not in ['csv', 'xlsx', 'xls']:
        flash('Formato inválido. Use CSV ou Excel (.xlsx, .xls).', 'error')
        return redirect(url_for('pedidos.listar_pedidos'))
    try:
        import pandas as pd
        conteudo = arquivo.read()
        if not conteudo:
            flash('Arquivo vazio.', 'error')
            return redirect(url_for('pedidos.listar_pedidos'))
        
        uploads_dir = Path(current_app.root_path).parent / 'uploads' / 'excel'
        uploads_dir.mkdir(parents=True, exist_ok=True)
        nome_limpo = arquivo.filename.replace(' ', '_')
        destino = uploads_dir / nome_limpo
        with open(destino, 'wb') as out_file:
            out_file.write(conteudo)
        
        df = pd.read_csv(BytesIO(conteudo)) if extensao == 'csv' else pd.read_excel(BytesIO(conteudo))
        if df.empty:
            flash('Arquivo sem linhas.', 'warning')
            return redirect(url_for('pedidos.listar_pedidos'))
        # Mapear colunas esperadas
        col_map = {
            'nome cliente': 'cliente_nome',
            'nome produto': 'produto_nome',
            'pedido id': 'pedido_id',
            'quantidade': 'quantidade',
            'preço venda': 'preco_venda',
            'preco venda': 'preco_venda',
            'data': 'data'
        }
        df.columns = [str(c).strip().lower() for c in df.columns]
        df = df.rename(columns=col_map)
        colunas_obrig = {'cliente_nome', 'produto_nome', 'quantidade', 'preco_venda', 'data', 'pedido_id'}
        if not colunas_obrig.issubset(set(df.columns)):
            flash('Colunas obrigatórias: Nome Cliente, Nome Produto, Pedido ID, Quantidade, Preço Venda, Data.', 'error')
            return redirect(url_for('pedidos.listar_pedidos'))
        
        # Garantir que os IDs de pedido formam a sequência esperada para evitar divergências
        pedido_ids_series = pd.to_numeric(df['pedido_id'], errors='coerce')
        if pedido_ids_series.isna().any():
            flash('Existem Pedido ID inválidos ou vazios na planilha. Corrija antes de importar.', 'error')
            return redirect(url_for('pedidos.listar_pedidos'))
        
        pedido_ids_unicos = sorted(set(int(x) for x in pedido_ids_series.tolist()))
        if not pedido_ids_unicos:
            flash('Planilha sem IDs de pedido válidos.', 'error')
            return redirect(url_for('pedidos.listar_pedidos'))
        
        total_pedidos_planilha = len(pedido_ids_unicos)
        colunas_para_processar = [c for c in ['cliente_nome','cliente_fantasia','produto_nome','quantidade','preco_venda','data','pedido_id'] if c in df.columns]
        # Resetar dados operacionais antes de importar o novo histórico
        PedidoService.resetar_dados_operacionais()
        resultado = PedidoService.processar_planilha_importacao(df[colunas_para_processar])
        resumo = resultado.get('resumo', {})
        erros_resultado = resultado.get('resultados', [])
        falhas = resumo.get('falha', 0)
        pedidos_criados = resumo.get('pedidos_criados', 0)
        
        if pedidos_criados != total_pedidos_planilha:
            flash(
                f"Atenção: a planilha possui {total_pedidos_planilha} pedidos únicos "
                f"mas o sistema criou {pedidos_criados}. Verifique a integridade do arquivo e tente novamente.",
                'warning'
            )
        elif pedidos_criados > 0:
            flash(f"{pedidos_criados} pedido(s) importado(s), espelhando o total da planilha.", 'success')

        if falhas > 0:
            exemplos = []
            for linha in erros_resultado[:3]:
                if linha.get('erros'):
                    exemplos.append(f"Linha {linha['linha']}: {linha['erros'][0]}")
            resumo_erros = ' | '.join(exemplos)
            if len(erros_resultado) > 3:
                resumo_erros = (resumo_erros + ' ...') if resumo_erros else '...'
            msg = f"{falhas} linha(s) precisam de correção manual."
            if resumo_erros:
                msg += f" Ex.: {resumo_erros}"
            flash(msg, 'warning')
            session['erros_importacao_pedidos'] = {
                'resumo': resumo,
                'resultados': erros_resultado
            }
            return redirect(url_for('pedidos.corrigir_importacao'))

        session.pop('erros_importacao_pedidos', None)
        return redirect(url_for('pedidos.listar_pedidos'))
    except ImportError:
        flash('Dependência ausente para processar planilhas.', 'error')
        return redirect(url_for('pedidos.listar_pedidos'))
    except Exception as e:
        current_app.logger.error(f"Erro ao importar histórico: {e}", exc_info=True)
        flash(f"Erro ao importar: {e}", 'error')
        return redirect(url_for('pedidos.listar_pedidos'))


@pedidos_bp.route('/importar/estornar', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def estornar_importacao():
    """Remove pedidos importados para permitir novo upload."""
    try:
        PedidoService.resetar_dados_operacionais()
        session.pop('erros_importacao_pedidos', None)
        
        uploads_dir = Path(current_app.root_path).parent / 'uploads' / 'excel'
        if uploads_dir.is_dir():
            arquivos_removidos = 0
            for arquivo in uploads_dir.iterdir():
                if arquivo.is_file():
                    try:
                        arquivo.unlink()
                        arquivos_removidos += 1
                    except Exception as rm_err:
                        current_app.logger.warning(f"Não foi possível remover arquivo {arquivo}: {rm_err}")
            if arquivos_removidos:
                current_app.logger.info(f"{arquivos_removidos} arquivo(s) de planilha removidos após estorno.")
        else:
            current_app.logger.info(f"Pasta de uploads não encontrada em {uploads_dir}")
        
        flash('Importação estornada. Todos os dados transacionais foram limpos.', 'success')
    except Exception as e:
        current_app.logger.error(f"Erro ao estornar importação: {e}", exc_info=True)
        flash('Erro ao estornar importação. Verifique os logs.', 'error')
    return redirect(url_for('pedidos.listar_pedidos'))


@pedidos_bp.route('/importar/corrigir', methods=['GET', 'POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def corrigir_importacao():
    """Exibe e processa correções manuais das linhas com erro."""
    if request.method == 'POST':
        try:
            linhas = request.form.getlist('linha')
            if not linhas:
                flash('Nenhuma linha enviada para correção.', 'warning')
                return redirect(url_for('pedidos.corrigir_importacao'))

            campos = {
                'cliente_nome': request.form.getlist('cliente_nome'),
                'cliente_fantasia': request.form.getlist('cliente_fantasia'),
                'produto_nome': request.form.getlist('produto_nome'),
                'quantidade': request.form.getlist('quantidade'),
                'preco_venda': request.form.getlist('preco_venda'),
                'data': request.form.getlist('data'),
                'pedido_id': request.form.getlist('pedido_id')
            }

            def obter_valor(lista, idx):
                if idx < len(lista):
                    valor = lista[idx]
                else:
                    valor = ''
                return valor.strip() if isinstance(valor, str) else valor

            registros = []
            for idx in range(len(linhas)):
                registros.append({
                    'cliente_nome': obter_valor(campos['cliente_nome'], idx),
                    'cliente_fantasia': obter_valor(campos['cliente_fantasia'], idx),
                    'produto_nome': obter_valor(campos['produto_nome'], idx),
                    'quantidade': obter_valor(campos['quantidade'], idx),
                    'preco_venda': obter_valor(campos['preco_venda'], idx),
                    'data': obter_valor(campos['data'], idx),
                    'pedido_id': obter_valor(campos['pedido_id'], idx)
                })

            import pandas as pd
            df_corrigido = pd.DataFrame(registros)
            resultado = PedidoService.processar_planilha_importacao(df_corrigido)
            resumo = resultado.get('resumo', {})
            falhas = resumo.get('falha', 0)
            pedidos_criados = resumo.get('pedidos_criados', 0)

            if pedidos_criados > 0:
                flash(f"{pedidos_criados} linha(s) corrigida(s) foram importadas.", 'success')

            if falhas > 0:
                session['erros_importacao_pedidos'] = {
                    'resumo': resumo,
                    'resultados': resultado.get('resultados', [])
                }
                flash('Ainda existem linhas com erro. Revise os campos destacados.', 'warning')
                return redirect(url_for('pedidos.corrigir_importacao'))

            session.pop('erros_importacao_pedidos', None)
            flash('Todas as linhas foram importadas com sucesso.', 'success')
            return redirect(url_for('pedidos.listar_pedidos'))

        except Exception as e:
            current_app.logger.error(f"Erro ao corrigir importação: {e}", exc_info=True)
            flash('Erro ao processar correções. Tente novamente.', 'error')
            return redirect(url_for('pedidos.corrigir_importacao'))

    dados = session.get('erros_importacao_pedidos')
    if not dados or not dados.get('resultados'):
        flash('Nenhuma importação pendente para correção.', 'info')
        return redirect(url_for('pedidos.listar_pedidos'))

    return render_template('corrigir_importacao_pedidos.html', resultado=dados)

@pedidos_bp.route('/novo', methods=['GET', 'POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def novo_pedido():
    """Cria um novo pedido"""
    if request.method == 'POST':
        # Extrair dados do formulário
        cliente_id = request.form.get('cliente_id')
        
        # Processar itens do pedido
        itens_data = []
        for produto_id, qtd, pv in zip(
            request.form.getlist('produto_id'),
            request.form.getlist('quantidade'),
            request.form.getlist('preco_venda')
        ):
            if produto_id and qtd and pv:
                # Limpar formatação do preço (R$ 32,00 -> 32.00)
                preco_limpo = pv.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                try:
                    preco_float = float(preco_limpo)
                    itens_data.append({
                        'produto_id': produto_id,
                        'quantidade': qtd,
                        'preco_venda': preco_float
                    })
                except ValueError:
                    current_app.logger.error(f"Erro ao converter preço: {pv}")
                    continue
        
        # Usar o serviço para criar o pedido
        sucesso, mensagem, pedido = PedidoService.criar_pedido(cliente_id, itens_data)
        
        if sucesso:
            flash(mensagem, 'success')
            return redirect(url_for('pedidos.listar_pedidos', filtro='todos'))
        else:
            flash(mensagem, 'error')
            # Retornar dados para o formulário em caso de erro
            clientes = Cliente.query.all()
            produtos = Produto.query.all()
            return render_template('novo_pedido.html', clientes=clientes, produtos=produtos)
    
    # GET: Mostrar formulário
    clientes = Cliente.query.all()
    produtos = Produto.query.all()
    return render_template('novo_pedido.html', clientes=clientes, produtos=produtos)

@pedidos_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def editar_pedido(id):
    """Edita um pedido existente"""
    pedido = PedidoService.buscar_pedido(id)
    if not pedido:
        flash('Pedido não encontrado', 'error')
        return redirect(url_for('pedidos.listar_pedidos'))

    if request.method == 'POST':
        # Extrair dados do formulário
        cliente_id = request.form.get('cliente_id')
        
        # Processar itens do pedido (similar ao novo_pedido)
        itens_data = []
        for produto_id, qtd, pv in zip(
            request.form.getlist('produto_id'),
            request.form.getlist('quantidade'),
            request.form.getlist('preco_venda')
        ):
            if produto_id and qtd and pv:
                preco_limpo = pv.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                try:
                    preco_float = float(preco_limpo)
                except ValueError:
                    current_app.logger.error(f"Erro ao converter preço: {pv}")
                    continue
                itens_data.append({
                    'produto_id': produto_id,
                    'quantidade': qtd,
                    'preco_venda': preco_float
                })
        
        # Usar o serviço para atualizar o pedido
        sucesso, mensagem, _pedido = PedidoService.editar_pedido(id, cliente_id, itens_data)
        
        if sucesso:
            current_app.logger.info(f"Pedido editado por {session.get('usuario_nome', 'N/A')}")
            flash(mensagem, 'success')
            return redirect(url_for('pedidos.listar_pedidos'))
        else:
            flash(mensagem, 'error')
            # Retornar dados para o formulário em caso de erro
            clientes = Cliente.query.all()
            produtos = Produto.query.all()
        return render_template('editar_pedido.html', pedido=pedido, clientes=clientes, produtos=produtos)
    
    clientes = Cliente.query.all()
    produtos = Produto.query.all()
    return render_template('editar_pedido.html', pedido=pedido, clientes=clientes, produtos=produtos)

@pedidos_bp.route('/confirmar/<int:id>', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def confirmar_pedido(id):
    """Confirma um pedido"""
    sucesso, mensagem = PedidoService.confirmar_pedido(id)
    
    if sucesso:
        flash(mensagem, 'success')
    else:
        flash(mensagem, 'error')
    
    return redirect(url_for('pedidos.listar_pedidos'))

@pedidos_bp.route('/editar/<int:id>/confirmar', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def confirmar_edicao_pedido(id):
    """Valida senha admin e libera edição de pedido."""
    try:
        senha_admin = request.form.get('senha')
        if not senha_admin:
            return jsonify({'success': False, 'message': 'Senha é obrigatória'}), 400

        # Validar senha do administrador
        from meu_app.models import Usuario
        admin = Usuario.query.filter_by(tipo='admin').first()
        if not admin or not admin.check_senha(senha_admin):
            return jsonify({'success': False, 'message': 'Senha incorreta'}), 403

        edit_url = url_for('pedidos.editar_pedido', id=id)
        return jsonify({'success': True, 'message': 'Senha validada. Redirecionando para edição.', 'redirect': edit_url})

    except Exception as e:
        current_app.logger.error(f"Erro ao confirmar edição do pedido: {str(e)}")
        return jsonify({'success': False, 'message': f'Erro ao salvar alterações: {str(e)}'}), 500

@pedidos_bp.route('/confirmar_comercial/<int:id>', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def confirmar_comercial_pedido(id):
    """Confirma pedido pelo comercial"""
    try:
        senha_admin = request.form.get('senha')
        
        if not senha_admin:
            return jsonify({'success': False, 'message': 'Senha é obrigatória'})
        
        # Validar senha do administrador
        usuario_logado = session.get('usuario_nome', 'N/A')
        current_app.logger.info(f"Tentativa de confirmação comercial do pedido {id} por {usuario_logado}")
        
        # Validar senha usando o serviço
        sucesso, mensagem = PedidoService.confirmar_pedido_comercial(id, senha_admin)
        
        if sucesso:
            current_app.logger.info(f"Pedido {id} confirmado comercialmente por {usuario_logado}")
            return jsonify({'success': True, 'message': mensagem})
        else:
            current_app.logger.warning(f"Falha na confirmação comercial do pedido {id}: {mensagem}")
            return jsonify({'success': False, 'message': mensagem})
            
    except Exception as e:
        current_app.logger.error(f"Erro ao confirmar pedido comercial {id}: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'})

@pedidos_bp.route('/excluir/<int:id>/confirmar', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def excluir_pedido_confirmar(id):
    """Exclui pedido após confirmação de senha"""
    try:
        senha_admin = request.form.get('senha')
        
        if not senha_admin:
            return jsonify({'success': False, 'message': 'Senha é obrigatória'})
        
        # Validar senha do administrador
        usuario_logado = session.get('usuario_nome', 'N/A')
        current_app.logger.info(f"Tentativa de exclusão do pedido {id} por {usuario_logado}")
        
        # Verificar senha do administrador
        from meu_app.models import Usuario
        admin = Usuario.query.filter_by(tipo='admin').first()
        if not admin or not admin.check_senha(senha_admin):
            return jsonify({'success': False, 'message': 'Senha incorreta'})
        
        # Excluir pedido
        sucesso, mensagem = PedidoService.excluir_pedido(id)
        
        if sucesso:
            current_app.logger.info(f"Pedido {id} excluído por {usuario_logado}")
            return jsonify({'success': True, 'message': mensagem})
        else:
            current_app.logger.warning(f"Falha na exclusão do pedido {id}: {mensagem}")
            return jsonify({'success': False, 'message': mensagem})
            
    except Exception as e:
        current_app.logger.error(f"Erro ao excluir pedido {id}: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro interno do servidor'})

@pedidos_bp.route('/cancelar/<int:id>', methods=['POST'])
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def cancelar_pedido(id):
    """Cancela um pedido"""
    sucesso, mensagem = PedidoService.cancelar_pedido(id)
    
    if sucesso:
        flash(mensagem, 'success')
    else:
        flash(mensagem, 'error')
    
    return redirect(url_for('pedidos.listar_pedidos'))

@pedidos_bp.route('/visualizar/<int:id>')
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def visualizar_pedido(id):
    """Visualiza os detalhes de um pedido"""
    try:
        pedido = PedidoService.buscar_pedido(id)
        if not pedido:
            flash("Pedido não encontrado", "error")
            return redirect(url_for("pedidos.listar_pedidos"))
    
        # Calcular totais usando o serviço
        totais = PedidoService.calcular_totais_pedido(id)
        
        # Capturar parâmetro origem para o template
        origem = request.args.get('origem', 'pedidos')
        retorno_params = {
            'filtro': request.args.get('filtro'),
            'data_inicio': request.args.get('data_inicio'),
            'data_fim': request.args.get('data_fim'),
            'sort': request.args.get('sort'),
            'direction': request.args.get('direction')
        }
        retorno_params = {chave: valor for chave, valor in retorno_params.items() if valor not in (None, '')}
        
        current_app.logger.info(f"Pedido {id} visualizado por {session.get('usuario_nome', 'N/A')} (origem: {origem})")
        
        return render_template('visualizar_pedido.html', 
                             pedido=pedido, 
                             total=totais['total'], 
                             pago=totais['pago'], 
                             saldo=totais['saldo'],
                             origem=origem,
                             retorno_params=retorno_params)
        
    except Exception as e:
        current_app.logger.error(f"Erro ao visualizar pedido: {str(e)}")
        flash("Erro ao carregar pedido", "error")
        return redirect(url_for("pedidos.listar_pedidos"))


@pedidos_bp.route('/importar/exemplo')
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def download_exemplo():
    """Baixa arquivo de exemplo para importação - formato Excel com melhor UX"""
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    try:
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exemplo Importação"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        info_font = Font(italic=True, color="666666")
        info_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Linha 1: Título
        ws['A1'] = "📋 EXEMPLO DE IMPORTAÇÃO DE PEDIDOS"
        ws['A1'].font = Font(bold=True, size=14, color="366092")
        ws.merge_cells('A1:F1')
        
        # Linha 2: Instruções
        ws['A2'] = "Instruções: Preencha os dados abaixo e salve como CSV para importar"
        ws['A2'].font = info_font
        ws.merge_cells('A2:F2')
        
        # Linha 3: Vazia
        ws['A3'] = ""
        
        # Linha 4: Cabeçalho das colunas
        headers = [
            "Cliente Nome",
            "Cliente Fantasia", 
            "Produto Nome",
            "Quantidade",
            "Preço Venda",
            "Data (YYYY-MM-DD)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Linha 5: Exemplo 1
        exemplo1 = [
            "CAIQUE ANDRADE NASCIMENTO",
            "KAIQUE ITATIM",
            "SKOL LATA 350ML",
            286,
            31.00,
            "2024-01-15"
        ]
        
        for col, value in enumerate(exemplo1, 1):
            cell = ws.cell(row=5, column=col, value=value)
            cell.border = border
            if col == 4 or col == 5:  # Quantidade e Preço
                cell.alignment = Alignment(horizontal="center")
        
        # Linha 6: Exemplo 2
        exemplo2 = [
            "CAIQUE ANDRADE NASCIMENTO",
            "KAIQUE ITATIM", 
            "BRAHMA CHOPP LATA 350 ML",
            286,
            32.00,
            "2024-01-15"
        ]
        
        for col, value in enumerate(exemplo2, 1):
            cell = ws.cell(row=6, column=col, value=value)
            cell.border = border
            if col == 4 or col == 5:  # Quantidade e Preço
                cell.alignment = Alignment(horizontal="center")
        
        # Linha 7: Exemplo 3
        exemplo3 = [
            "LUCIANO VIEIRA SILVA DE ARAUJO",
            "LUCIANO MOURA",
            "RED BULL ENERGY DRINK 250 ML",
            144,
            150.00,
            "2024-01-16"
        ]
        
        for col, value in enumerate(exemplo3, 1):
            cell = ws.cell(row=7, column=col, value=value)
            cell.border = border
            if col == 4 or col == 5:  # Quantidade e Preço
                cell.alignment = Alignment(horizontal="center")
        
        # Linha 8: Vazia
        ws['A8'] = ""
        
        # Linha 9: Dicas importantes
        ws['A9'] = "💡 DICAS IMPORTANTES:"
        ws['A9'].font = Font(bold=True, color="366092")
        ws.merge_cells('A9:F9')
        
        # Linha 10: Dica 1
        ws['A10'] = "• Data deve estar no formato: YYYY-MM-DD (ex: 2024-01-15)"
        ws['A10'].font = info_font
        ws.merge_cells('A10:F10')
        
        # Linha 11: Dica 2
        ws['A11'] = "• Preços devem usar ponto como separador decimal: 31.00 (não 31,00)"
        ws['A11'].font = info_font
        ws.merge_cells('A11:F11')
        
        # Linha 12: Dica 3
        ws['A12'] = "• Nomes de clientes e produtos devem existir no sistema"
        ws['A12'].font = info_font
        ws.merge_cells('A12:F12')
        
        # Linha 13: Dica 4
        ws['A13'] = "• Salve este arquivo como CSV (.csv) antes de importar"
        ws['A13'].font = info_font
        ws.merge_cells('A13:F13')
        
        # Ajustar largura das colunas
        column_widths = [25, 20, 30, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name='exemplo_importacao_pedidos.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        current_app.logger.error(f"Erro ao gerar arquivo de exemplo: {str(e)}")
        flash('Erro ao gerar arquivo de exemplo', 'error')
        return redirect(url_for('pedidos.importar_pedidos'))

# Modelo simplificado (.xlsx) para importação em massa
@pedidos_bp.route('/download_modelo_pedidos')
@login_obrigatorio
@permissao_necessaria('acesso_pedidos')
def download_modelo_pedidos():
    """Gera planilha-modelo com cabeçalhos solicitados e data em padrão brasileiro."""
    from datetime import datetime
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Modelo Pedidos'

    headers = [
        'Nome Cliente',
        'Nome Produto',
        'Pedido ID',
        'Quantidade',
        'Preço Venda',
        'Data'
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    exemplo = [
        'Ex: João da Silva',
        'Ex: SKOL LATA 350ML',
        'Ex: 12345',
        10,
        25.50,
        datetime(2024, 1, 15)
    ]

    for col, value in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = border
        if col in (4, 5):
            cell.alignment = Alignment(horizontal='center')
            if col == 5:
                cell.number_format = 'R$ #.##0,00'
        if col == 6:
            cell.number_format = 'DD/MM/YYYY'

    ws['A3'] = ''
    ws['A4'] = 'Observações:'
    ws['A4'].font = Font(bold=True, color='366092')
    ws.merge_cells('A4:F4')
    ws['A5'] = '• Preencha SEMPRE o Pedido ID (mesmo identificador usado no sistema antigo).'
    ws.merge_cells('A5:F5')
    ws['A6'] = '• Informe a data no formato DD/MM/AAAA (Padrão Brasil) para coincidir com o novo layout.'
    ws.merge_cells('A6:F6')
    ws['A7'] = '• Informe o preço com vírgula como separador decimal (ex.: 25,50).'
    ws.merge_cells('A7:F7')

    column_widths = [25, 30, 18, 15, 15, 18]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='modelo_pedidos.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
